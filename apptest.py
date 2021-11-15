import tkinter as tk
from tkinter import ttk
import openpyxl 
import os
import datetime
from fuzzywuzzy import fuzz

    
def schpQueryProcess(): 
    
    #os.chdir(r"C:\Users\jones.2541\Desktop")
    #resultswb = openpyxl.Workbook()
    scholarshipPSTD = ''
    unappliedwb = ''
    enrollmentHours = ''
    ##resultswb.save('newfile.xlsx')
    ##print(datetime.date.today())
    os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries")
    fmax = max(os.listdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries")[:-1]) #finds most recent file; exludes 'query results' adds file names to list alphabetically 
    #for i,e in enumerate(os.listdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries\FY22") ): ##need to account for folder change, find most recent date modified and name starting with 'FY'
    for i,e in enumerate(os.listdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries" + "\\"+ str(fmax))):
        #print(e[0:10])
        #print(datetime.datetime.strptime(e[0:10],"%Y_%m_%d").date() ,datetime.date.today())
        #print(os.getcwd())
        #print(datetime.date.today(),datetime.datetime.strptime(e[0:10],"%Y_%m_%d").date())
        
        if datetime.date.today() == datetime.datetime.strptime(e[0:10],"%Y_%m_%d").date():
            #os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries\FY22") ##need to account for folder change
            os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries" + "\\"+ str(fmax))
            #print('true', e[0:10])
            #print(e[11:])
            if e[11:] == 'OSF_SCHOLARSHIP_ENROLMNT_HOURS.xlsx':
                enrollmentHours = openpyxl.load_workbook(e)
                #print(e,type(e), os.getcwd()) checking type and directory, was still in the default desktop location
            elif e[11:] == 'OSF_SCHOLARSHIP_PSTD_ENROLMNT.xlsx':
                scholarshipPSTD = openpyxl.load_workbook(e)
            elif e[11:] == 'OSF_UNAPPLIED_CREDITS_FILTER.xlsx':
                unappliedwb = openpyxl.load_workbook(e)
         
    #os.chdir(r"C:\Users\jones.2541\Desktop")
    

    sheet = scholarshipPSTD.active
    ScholarshipItemTypes = ['050000000014','050000000016','050000000019','050000000022']
    enrollmentHourSheet = enrollmentHours.active
    unappliedSheet = unappliedwb.active
    


    ####vlookup####
    os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Exceptions")
    emax = max(os.listdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Exceptions"))
    #exceptions = openpyxl.load_workbook('1218 Exceptions.xlsx') ##need to account for file change, find most recent date modified
    exceptions = openpyxl.load_workbook(emax)
    execsheet = exceptions.active
    vlookup = set()
    for i in list(execsheet.columns)[0]:
        vlookup.add(str(i.value))

    ###creating a new spreadsheet to save results
    resultswb = openpyxl.Workbook()
    resultsSheet = resultswb.active #sheet 1
    resultsSheet.title = 'PSTD Results'
    resultswb.create_sheet(index=1, title= 'Unapplied Results')
    resultswb.create_sheet(index=2, title= 'Enrollment Hour Results')
    resultsSheet2 = resultswb['Unapplied Results']
    resultsSheet3 = resultswb['Enrollment Hour Results']



    headerList = ['ID','Item Type','Descr','Item Amt','Term','Take Prgrs','Career','Ref Nbr','Postd DtTm','User']
    for i, ele in enumerate(headerList):
        resultsSheet.cell(row=1, column= i+1).value = ele
        resultsSheet2.cell(row=1, column= i+1).value = ele
        resultsSheet3.cell(row=1, column= i+1).value = ele

    resultsSheet2.cell(row=1, column=11).value = 'Take Prgrs' #orginal spreadsheet has different headers


    ###PSTD Query####
    #print('from pstd query')
    count = 0

    a = list(sheet.columns)[5] #list of column 'Take Prgrs'/credit hours of each student
    for i in range(len(a)): #loops through each element and checks to see if it is equal to 0 
        if a[i].value == 0: #problem here with iknowican as it will pull in the students < fulltime that are returned to donor;; add comparsion to exceptions list?
            #print(sheet.cell(row=i+1, column=1).value)
            if sheet.cell(row=i+1, column=1).value not in vlookup: 
                count += 1
                #print(sheet.cell(row=i+1, column=1).value)
                for index, ele in enumerate(list(sheet.rows)[i]): #if element in 'a' is 0, grabs every element in row i
                    #print(ele.value, end=" ")
                    resultsSheet.cell(row= count+1, column=index+1).value = ele.value
                #print('\n')
                
    ###IKIC Logic###
    ikicList = [
    'I KNOW I CAN',
    'IKIC',
    ]
    #needs to check exceptions list
    for listitem in ikicList:
        for i, ele in enumerate(list(sheet.columns)[7]):
            if fuzz.token_set_ratio(listitem, ele.value) > 90 and int(sheet.cell(row=i+1, column=6).value) < 12 and sheet.cell(row=i+1, column=1).value not in vlookup:
               count += 1
               for j, elej in enumerate(list(sheet.rows)[i]):
                    #print(elej.value, end=" ")
                    resultsSheet.cell(row= count+1, column=j+1).value = elej.value
               #print('\n')    
    #print('Count = ', count, datetime.date.today())
    if count == 0:
        resultsSheet.cell(row= count+2, column=1).value = 'No Results'


    ####unapplied query####
    unappliedCount = 0
    for index, i in enumerate(list(unappliedSheet.columns)[3]):
        if i.value in ScholarshipItemTypes and unappliedSheet.cell(row=index+1, column=11).value == 0 :
            #print(i.value, index, unappliedSheet.cell(row=index+1, column=1).value)
            unappliedCount += 1
            for indexj, ele in enumerate(list(unappliedSheet.rows)[index]):
                #print(ele.value, end=" ")  
                resultsSheet2.cell(row = unappliedCount+1, column =indexj+1).value = ele.value
                #adds values to 1st row and each adjacent column in new spreadsheet
            #print('\n')
           

    if unappliedCount == 0:
        resultsSheet2.cell(row = unappliedCount+2, column =1).value = 'No Results'
    """need to compare to exceptions list """

    ####enrollment hour query####
    hItemTypes = ['050000000033',
    '050000000035',
    '050000000021',
    '050000000023'
    ]
    enrollmentCount = 0
    for i, ele in enumerate(list(enrollmentHourSheet.columns)[5]):
        #print(ele.value)
        if ele.value == None or type(ele.value) == str:
            continue
        elif enrollmentHourSheet.cell(row=i+1, column= 2).value in hItemTypes and ele.value > 0:
            enrollmentCount += 1
            for j, elej in enumerate(list(enrollmentHourSheet.rows)[i]):
                #print(elej.value, end=" ")
                resultsSheet3.cell(row= enrollmentCount+1, column=j+1).value = elej.value
            #print('\n')

    if enrollmentCount == 0:
        resultsSheet3.cell(row= enrollmentCount+2, column = 1).value = "No Results"
    
    
    os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries\Query Results")
    #os.chdir(r"C:\Users\jones.2541\Desktop")
    current_date = datetime.date.today()
    resultswb.save('Query Results_' + str(current_date) + '.xlsx')  

    ##########
    ##ALT LOAN PROCESS###
    ##########
    os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Training")

    wb = openpyxl.load_workbook('List of commonly received Alternative Loans.xlsx')
    sheet = wb.active


    a = list(sheet.columns)[0]
    altloanNames = []
    for i in range(1,len(a)):
        altloanNames.append(a[i].value)

    #os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries\FY22") ##need to account for folder change
    os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries" + "\\"+ str(fmax))
    schpwb = scholarshipPSTD
    #schpwb = openpyxl.load_workbook('_Altloan_Testfile_2020_11_04_OSF_SCHOLARSHIP_PSTD_ENROLMNT.xlsx')
    schpsheet = schpwb.active

    refcolumn = list(schpsheet.columns)[7]

    os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Alt Loan Query Results")
    altLoanresults = openpyxl.Workbook()
    resultsSheet = altLoanresults.active
    resultsSheet.cell(row=1, column=1).value = 'ID'
    resultsSheet.cell(row=1, column=2).value = 'Item Type'
    resultsSheet.cell(row=1, column=3).value = 'Descr'
    resultsSheet.cell(row=1, column=4).value = 'Item Amt'
    resultsSheet.cell(row=1, column=5).value = 'Term'
    resultsSheet.cell(row=1, column=6).value = 'Take Prgrs'
    resultsSheet.cell(row=1, column=7).value = 'Career'
    resultsSheet.cell(row=1, column=8).value = 'Ref Nbr'
    resultsSheet.cell(row=1, column=9).value = 'Postd DtTm'
    resultsSheet.cell(row=1, column=10).value = 'User'

    count = 0
    if datetime.datetime.today().weekday() == 0:
        for i in altloanNames:
            for j, ele in enumerate(refcolumn):
                if schpsheet.cell(row=j+1, column=9).value == None or type(schpsheet.cell(row=j+1, column=9).value) == str: #moves on if column has nothing in it or a string
                    continue
                #elif schpsheet.cell(row=j, column=9).value.date() == datetime.date.today()-datetime.timedelta(1):
                elif schpsheet.cell(row=j+1, column=9).value.date() == datetime.date.today()-datetime.timedelta(3) and fuzz.token_set_ratio(i, ele.value) > 90:
                    #print(fuzz.token_set_ratio(i, ele.value),i ,ele.value)
                    count += 1
                    for index, element in enumerate(list(schpsheet.rows)[j]): #why not j+1 here?
                        resultsSheet.cell(row=count+1, column=index+1).value = element.value
                        #print(element.value, end=" ")
                    #print('\n')
    else:  
        for i in altloanNames:
            for j, ele in enumerate(refcolumn):
                if schpsheet.cell(row=j+1, column=9).value == None or type(schpsheet.cell(row=j+1, column=9).value) == str: #moves on if column has nothing in it or a string
                    continue
                #elif schpsheet.cell(row=j, column=9).value.date() == datetime.date.today()-datetime.timedelta(1):
                #else:
                #   print(fuzz.token_set_ratio(i, ele.value),i,ele.value, schpsheet.cell(row=j+1, column=1).value) 
                elif schpsheet.cell(row=j+1, column=9).value.date() == datetime.date.today()-datetime.timedelta(1) and fuzz.token_set_ratio(i, ele.value) > 90:
                    #print(fuzz.token_set_ratio(i, ele.value),i, ele.value)
                    count += 1
                    for index, element in enumerate(list(schpsheet.rows)[j]): #why not j+1 here?
                        resultsSheet.cell(row=count+1, column=index+1).value = element.value
                        #print(element.value, end=" ")
                    #print('\n')
    altLoanresults.save('Alt Loan Results_'+ str(datetime.date.today()) + '.xlsx')
    runCompletedWindow() #pop up to say completed
    

def runCompletedWindow():
    window = tk.Toplevel()
    window.title('Completed Window')
    window.geometry("200x100")
    label = ttk.Label(window,text = "Process Complete")
    #window.configure(bg='green')
    
    topLevelButton = tk.Button(window,
                          text = "CLOSE", 
                          fg="white",
                          bg="red",
                          relief = 'solid',
                          command=window.destroy)
    #button.grid(row=1,column=0)
    label.pack()
    topLevelButton.pack()
    
    
####################
root = tk.Tk()
root.title("External Scholarship Hold Process")
style = ttk.Style()
style.configure('TFrame', background='grey')#area behind buttons

frame = ttk.Frame(root)
#frame.pack()
frame.grid()
#frame.place()
root.geometry("400x300")
root.columnconfigure(0, weight=1)   # Set weight to row and 
root.rowconfigure(0, weight=1)      # column where the widget is
root.configure(bg='grey') #area behind everything

#img = tk.PhotoImage(file = r"C:\Users\jones.2541\Desktop\button.png") add an image

#####################

schpQPbutton = tk.Button(frame, 
                text = "Click to run Scholarship Query Process and Alt Loan Process", 
                command = schpQueryProcess,
                height = 8,
                relief = 'solid'
                
                )
schpQPbutton.grid(row=0,column=0)
#schpQPbutton.pack(side=tk.TOP, pady=20, padx=8)
#schpQPbutton.place(x=200, y=150)

button = tk.Button(frame, 
                   text="QUIT", 
                   fg="white",
                   bg="red",
                   relief = 'solid',
                   command=root.destroy)
                   
#button.pack(side=tk.BOTTOM, pady=8, padx=8)
button.grid(row=1,column=0, pady= 20)




root.mainloop()
